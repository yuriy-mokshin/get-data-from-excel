from openpyxl import  load_workbook
from objects.customer import Customer
from objects.customer import Counter


def get_data_from_excel():
    file_excel_1 = "Book2.xlsx"
    file_excel_2 = "Book1.xlsx"
    customer_list = []
    # Processing the first document
    book1 = load_workbook(filename=file_excel_1)
    sheet = book1['Sheet1']
    varee_id = ''
    for index in range(2, 56):
        varee_id = sheet['C' + str(index)].value
        print(f"[+] - step: {str(index)}. varee_id = {varee_id}")
        customer = Customer(sheet['D' + str(index)].value,
                            varee_id,
                            sheet['AA' + str(index)].value,
                            sheet['N' + str(index)].value,
                            sheet['O' + str(index)].value)
        customer.short_salutation = sheet['M' + str(index)].value
        customer.title = ''
        customer.set_address('', '', sheet['Q' + str(index)].value, sheet['R' + str(index)].value, '', sheet['S' + str(index)].value, sheet['T' + str(index)].value)

        # *** Checking dubilkates for varee_id ***
        is_new = True
        for x in customer_list:
            if varee_id == x.contract:
                is_new = False
                break
        if is_new is True:
            customer_list.append(customer)
        # *** ***

    # Processing the second document
    book2 = load_workbook(filename=file_excel_2)
    sheet2 = book2['Sheet1']
    for i in range(0, len(customer_list)):
        for index in range(1, 56):
            varee_id = sheet2['F' + str(index)].value
            if customer_list[i].contract == varee_id:
                customer_list[i].add_counter(sheet2['AL' + str(index)].value,
                                             sheet2['AM' + str(index)].value,
                                             sheet2['AO' + str(index)].value,
                                             sheet2['AN' + str(index)].value,
                                             sheet2['AS' + str(index)].value,
                                             sheet2['AT' + str(index)].value,
                                             sheet2['AV' + str(index)].value)



    # *** log ***
    print(f"[+]: Customers count: {str(len(customer_list))}")
    for customer in customer_list:
        print(f"\n[+]: PIN: {customer.pin} | contract: {customer.contract} | pwd: {customer.password} | fname: {customer.first_name} | lname: {customer.last_name} | city: {customer.community} | street: {customer.street} | h_n: {customer.house_number}")

        print(f"[+]: Counters count: {str(len(customer.get_counters))}")
        for counter in customer.get_counters:
            print(f" - id: {counter.id}, number: {counter.number} | obis_code: {counter.obis_code} | obis_code_short_name: {counter.obis_code_short_name} | min_value: {counter.min_value} | max_value: {counter.max_value} | last_value: {counter.last_value}")
    # *** ***
    return customer_list


def main():
    customer_list = get_data_from_excel()


if __name__ == "__main__":
    main()
